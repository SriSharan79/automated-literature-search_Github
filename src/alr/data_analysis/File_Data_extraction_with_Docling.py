from alr.common.file_manager import DataAnalyzeManager
from alr.data_analysis.Data_analysis_system_prompts import SYSTEM_PROMPT_Body_Identifier, SYSTEM_PROMPT_Heading_identifier
from alr.data_analysis.Data_sorting_utils import categorize_sections, merge_content_by_refined_headings, merge_content_by_refined_headings_Chunks, process_document_chunks, save_excel_log, sort_raw_text_with_headings
from alr.common.general_utils import _as_path, caluculate_time_taken
from alr.data_analysis.LLM_Reference_Extractor import process_references_from_chunks
from alr.common.llm_utils import llm_call
from alr.data_analysis.Refrences_log_utils import log_Ref_data_extracted
from alr.data_analysis.Table_image_extractor import DoclingExtractor


import re
from docling.document_converter import DocumentConverter
from docling.chunking import HybridChunker
from colorama import Fore,Style

import time  # <--- Add this line
import json
import traceback
from datetime import datetime
from pathlib import Path
import os
import pandas as pd
import json
import ast
from openpyxl import load_workbook
import traceback


converter = DocumentConverter()
chunker = HybridChunker()


def process_llm_refined_structure(llm_output_string):
    # This regex finds the first occurrence of '[' and the last ']' 
    # and extracts everything in between, including the brackets.
    match = re.search(r'(\[.*\])', llm_output_string, re.DOTALL)
    
    if match:
        cleaned_input = match.group(1).strip()
    else:
        # Fallback to the original logic if no brackets are found
        cleaned_input = llm_output_string.strip()

    try:
        # ast.literal_eval is great because it's safer than eval()
        return ast.literal_eval(cleaned_input)
    except (ValueError, SyntaxError) as e:
        print(f"Error parsing: {e}")

        traceback.print_exc()
        return []

def get_llm_refined_lists(unique_headings_list,llm_service):
    """Handles the two-step LLM refinement process."""
    # Step 1: Get Refined List
    prompt_refined = f"List of Headings: {unique_headings_list}"
    res_refined = llm_call(prompt_refined, SYSTEM_PROMPT_Heading_identifier,llm_service)
    text_refined = res_refined.text if hasattr(res_refined, 'text') else str(res_refined)
    refined_headings = process_llm_refined_structure(text_refined)


    # print(Fore.GREEN + f"\n refined_headings Success: {refined_headings}." + Fore.RESET)

    # Step 2: Get Body Headings
    prompt_body = f"List of Headings: {refined_headings}"
    res_body = llm_call(prompt_body, SYSTEM_PROMPT_Body_Identifier,llm_service)
    text_body = res_body.text if hasattr(res_body, 'text') else str(res_body)
    body_headings = process_llm_refined_structure(text_body)


    # print(Fore.GREEN + f"\n body_headings Success: {body_headings}." + Fore.RESET)

    return refined_headings, body_headings


def reference_pipeline_worker(chunks, doc, ref_json_path, excel_log_path, pdf_name, ID):
    """
    Helper function to run the extraction and logging sequentially 
    inside a background thread.
    """
    try:
        # 1. First, extract the references (This takes time/LLM calls)
        process_references_from_chunks(chunks, doc, ref_json_path)

        # 2. Safety Buffer: Wait up to 5 seconds for the file to actually appear on disk
        attempts = 0
        while not os.path.exists(ref_json_path) and attempts < 5:
            time.sleep(1)
            attempts += 1

        # 3. Logging
        log_Ref_data_extracted(excel_log_path, ref_json_path, pdf_name, ID)
        
        # print(Fore.CYAN + f"\n[Background] Reference processing complete for {pdf_name}" + Fore.RESET)
    except Exception as e:
        print(Fore.RED + f"\n[Background Error] Reference Pipeline failed: {e}" + Fore.RESET)
        traceback.print_exc()

def _init_excel_data(file_path, ID):
    pdf_name = Path(file_path).name
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    excel_data = {
        "file name": pdf_name,
        "file path": str(file_path),
        "Time Stamp": current_time,
        "UUID": ID,
        "Chunking Time": "Time record Failed",
        "Sectioning Time": "Time record Failed",
        "output file name": f"{ID}_sections.json",
        "Abstract": "No",
        "Keywords": "No",
        "Introduction": "No",
        "Conclusion": "No",
        "References": "No",
        "Acknowledgments": "No",
        "Body": [],
        "Remaining": [],
    }
    return pdf_name, excel_data


def _excel_log_has_file(excel_path, pdf_name, filename_col_candidates=None):
    """
    Returns True if excel_path exists and contains a row whose 'file name' column matches pdf_name.
    Works even if the column casing differs. If the column isn't found, returns False (fail-open).
    """
    if filename_col_candidates is None:
        filename_col_candidates = ["file name", "filename", "pdf name", "file", "document", "name"]

    try:
        excel_path = Path(excel_path)
        if not excel_path.exists():
            return False

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        header = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            header.append(str(cell).strip() if cell is not None else "")

        # Find best matching filename column
        header_lower = [h.lower() for h in header]
        col_idx = None
        for cand in filename_col_candidates:
            if cand.lower() in header_lower:
                col_idx = header_lower.index(cand.lower()) + 1  # 1-based index
                break

        if col_idx is None:
            return False

        # Scan column values
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx - 1]
            if val is None:
                continue
            if str(val).strip() == str(pdf_name).strip():
                return True

        return False

    except Exception as e:
        # Fail-open: if checking the log fails, don't skip processing
        print(Fore.YELLOW + f"⚠ Error checking Excel log '{excel_path}': {e}" + Fore.RESET)
        traceback.print_exc()
        return False


def _decide_skip_paths(Main_Folder, pdf_name):
    """
    Checks both raw_section log and reference Excel log for this pdf_name.
    - If section log has entry -> skip section pipeline + section logging
    - If reference log has entry -> skip reference pipeline + reference logging
    - If both exist -> do "abstract only" (no section/ref logging)
    """
    section_exists = _excel_log_has_file(Main_Folder.raw_section_excel_log_path, pdf_name)
    ref_exists = _excel_log_has_file(Main_Folder.refrences_excel_log_path, pdf_name)

    if section_exists and ref_exists:
        print(Fore.WHITE + f"ℹ Info: refrence data '{pdf_name}' already exists . Running ABSTRACT-ONLY flow." + Fore.RESET)
        return True, True, True  # skip_sections, skip_refs, abstract_only

    if section_exists:
        print(Fore.WHITE + f"ℹ Info: text processed data for'{pdf_name}' already exists. Skipping text processing and logging." + Fore.RESET)

    if ref_exists:
        print(Fore.WHITE + f"ℹ Info: refrence data '{pdf_name}' already exists in  Skipping reference extraction + reference logging." + Fore.RESET)

    return section_exists, ref_exists, True


def _extract_and_chunk(MF,logger, file_path, start_time):
    CHUNKS_CACHE_FILE = MF.raw_chunks_json_path
    
    chunks_data = []   
    doc = None 
    tables_data = []
    images_data = []
    headings_data = []
    
    # 1. Attempt to load from cache
    if os.path.exists(CHUNKS_CACHE_FILE):
        try:
            logger.info(f"Attempting to load data from cache file: {CHUNKS_CACHE_FILE}")
            with open(CHUNKS_CACHE_FILE, 'r', encoding='utf-8') as cache_file:
                cached_content = json.load(cache_file)
                
                # Check if cache is in the new dictionary format or older list format
                if isinstance(cached_content, dict) and "chunks" in cached_content:
                    chunks_data = cached_content["chunks"]
                    tables_data = cached_content.get("tables", [])
                    images_data = cached_content.get("images", [])
                    headings_data = cached_content.get("headings", [])
                    logger.info("Successfully loaded chunks, tables, and images from structured JSON cache.")
                else:
                    chunks_data = cached_content
                    logger.info("Successfully loaded chunks from legacy list JSON cache.")
                    
            print(Fore.GREEN + "ℹ Info: Loaded chunks successfully from JSON cache file." + Fore.RESET)
        except Exception as e:
            print(Fore.YELLOW + f"Failed to parse cache file: {e}" + Fore.RESET)
            logger.error(f"Failed to parse cache file: {e}", exc_info=True)
            traceback.print_exc()
    
    # 2. Cache miss: Run processing pipelines
    if not chunks_data:
        print(Fore.YELLOW + "Cache missing. Running backend conversion pipelines via Docling...")
        logger.info("Cache miss. Initializing DoclingExtractor for pipeline extraction.")
        
        try:
            # Initialize DoclingExtractor using storage paths from MF configuration
            extractor = DoclingExtractor(
                input_path=file_path,
                tables_output_path=MF.tables_storage_path,
                images_output_path=MF.image_storage_path
            )
            
            # Convert file exactly ONCE using the Extractor's pipeline configuration 
            # (Ensures pipeline options like 'generate_picture_images' are active)
            logger.info(f"Starting single-pass Docling document conversion for: {file_path}")
            conversion_result = extractor.doc_converter.convert(file_path)
            doc = conversion_result.document
            
            # Extract layout artifacts directly from the unified conversion object
            logger.info("Extracting tables to storage path...")
            tables_data = extractor._extract_tables(conversion_result)
            
            logger.info("Extracting images to storage path with deduplication...")
            images_data = extractor._extract_images(conversion_result)
            
            logger.info("Extracting layout headings...")
            headings_data = extractor._extract_headings(conversion_result)
            
            # Extract original text chunks via the provided chunker
            logger.info("Executing text chunker pipeline...")
            raw_chunks = list(chunker.chunk(dl_doc=doc))
            
            # Transform text chunks into a JSON-serializable structure
            for chunk in raw_chunks:
                chunk_pages = set()
                chunk_dict = {
                    "text": chunk.text,
                    "meta": {
                        "headings": chunk.meta.headings if hasattr(chunk.meta, 'headings') else [],
                        "doc_items": []
                    }
                }
                
                # Capture doc_items metadata, underlying text contents, and page numbers
                if hasattr(chunk.meta, 'doc_items'):
                    for item in chunk.meta.doc_items:
                        item_text = None
                        self_ref = getattr(item, 'self_ref', None)
                        if self_ref:
                            match = re.match(r'#/texts/(\d+)', self_ref)
                            if match:
                                text_index = int(match.group(1))
                                try:
                                    item_text = doc.texts[text_index].text
                                except (IndexError, AttributeError):
                                    item_text = None

                        # Extract page_no from the provenance (prov) list
                        item_pages = []
                        prov_list = getattr(item, 'prov', [])
                        if prov_list:
                            for prov_item in prov_list:
                                p_no = prov_item.get("page_no") if isinstance(prov_item, dict) else getattr(prov_item, "page_no", None)
                                if p_no is not None:
                                    item_pages.append(p_no)
                                    chunk_pages.add(p_no)

                        item_dict = {
                            "label": getattr(item, 'label', 'N/A'),
                            "self_ref": self_ref,
                            "actual_text": item_text,
                            "page_no": item_pages[0] if item_pages else None,   # Primary/first page for this item
                            "all_pages": item_pages                             # Full list of pages if it spans multiple
                        }
                        chunk_dict["meta"]["doc_items"].append(item_dict)
                
                # Log a consolidated sorted list of page numbers for the entire chunk
                chunk_dict["meta"]["page_numbers"] = sorted(list(chunk_pages))
                chunks_data.append(chunk_dict)

            print(Fore.GREEN + f"ℹ Info: Chunking done" + Fore.RESET)
            logger.info(f"Chunking process complete. Extracted {len(chunks_data)} chunks.")
            
            # Save structural representations, tables, and images into JSON cache safely
            cache_payload = {
                "chunks": chunks_data,
                "tables": tables_data,
                "images": images_data,
                "headings": headings_data
            }
            
            logger.info(f"Writing structured extraction payload to cache file: {CHUNKS_CACHE_FILE}")
            with open(CHUNKS_CACHE_FILE, 'w', encoding='utf-8') as cache_file:
                json.dump(cache_payload, cache_file, indent=4, ensure_ascii=False)
            print(Fore.GREEN + "Backend parsing complete. Structure saved to cache.")
            
        except Exception as e:
            print(Fore.RED + f"Critical Exception during Docling processing: {e}" + Fore.RESET)
            logger.error(f"Critical Exception during Docling processing: {e}", exc_info=True)
            traceback.print_exc()

    chunking_end_time = time.time()
    time_taken_4_chunking = caluculate_time_taken(start_time, chunking_end_time)
    
    logger.info(f"Extraction and chunking workflow wrapped execution in {time_taken_4_chunking}.")
    return doc, chunks_data, chunking_end_time, time_taken_4_chunking

def _process_sections(chunks, doc,llm_service):
    Section_texts, unique_headings = process_document_chunks(chunks, doc)
    Sections_Raw_Chunk_text = sort_raw_text_with_headings(chunks, doc)
    refined_headings, body_headings = get_llm_refined_lists(unique_headings,llm_service)

    final_merged_content = merge_content_by_refined_headings(Section_texts, refined_headings)
    refined_Sections_Raw_chunks = merge_content_by_refined_headings_Chunks(
        Sections_Raw_Chunk_text, refined_headings
    )

    return final_merged_content, refined_Sections_Raw_chunks, body_headings


def _categorise_sections(refined_Sections_Raw_chunks, final_merged_content, body_headings, excel_data):
    processed_data_for_json, excel_data = categorize_sections(
        refined_Sections_Raw_chunks,
        final_merged_content,
        body_headings,
        excel_data,
    )
    return processed_data_for_json, excel_data


def _save_section_outputs(processed_data_for_json, Main_Folder, excel_data, time_taken_4_chunking, time_taken_4_section_processing, pdf_name):
    section_json_path = Main_Folder.raw_sec_json_path

    with open(section_json_path, "w", encoding="utf-8") as f:
        json.dump(processed_data_for_json, f, indent=4)

    save_excel_log(
        Main_Folder.raw_section_excel_log_path,
        excel_data,
        time_taken_4_chunking,
        time_taken_4_section_processing,
    )
    print(Fore.GREEN + f"✓ Success: {pdf_name} sections logged." + Fore.RESET)



if __name__ == "__main__":

    folder_path="/localdata/user/kata_du/Automated Literature Survey/02_Test_Storage"

    Main_Folder = DataAnalyzeManager(folder_path)
    pdf_name="2022-Review Model-based Systems Engineering and Artific.pdf"
    ID="26cff1e1"

    Main_Folder.update_id_files(ID)

    ref_json_path = _as_path(Main_Folder.ref_json_path)
    log_Ref_data_extracted(
            Main_Folder.refrences_excel_log_path,
            ref_json_path,
            pdf_name,
            ID
        )
    # Main_Folder.update_id_files("test_id")

    # file_path="/localdata/user/kata_du/Automated Literature Survey/downloads/MBSE and AI/2022-Review Model-based Systems Engineering and Artific.pdf"

    # docling_process_file(file_path,Main_Folder.current_id,Main_Folder)