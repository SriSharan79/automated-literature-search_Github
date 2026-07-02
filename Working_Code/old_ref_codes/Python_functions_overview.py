#!/usr/bin/env python3
"""
Map Python function definitions and where they are used (call-sites) across a folder,
then export to a single Excel (.xlsx) with two sheets: Definitions + Usages.

Usage:
  python map_defs_and_usages_to_excel.py /path/to/folder -o function_map.xlsx

Dependency:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import ast
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ----------------------------
# Data models
# ----------------------------

@dataclass(frozen=True)
class DefKey:
    """Unique-ish identifier for a definition."""
    module: str              # e.g. package.subpkg.file
    qualname: str            # e.g. foo, MyClass.method, outer.inner
    def_type: str            # function/method/nested/async_...


@dataclass
class FunctionDefRecord:
    key: DefKey
    file_path: str
    module: str
    scope: str               # human readable scope: <module>, MyClass, outer, MyClass.outer, ...
    name: str                # function/method name
    qualname: str            # dotted path
    def_type: str
    signature: str
    decorators: str
    lineno: int
    end_lineno: int


@dataclass
class UsageRecord:
    used_in_file: str
    used_in_module: str
    used_in_lineno: int
    call_text: str                  # reconstructed best-effort, e.g. foo(...), mod.foo(...), obj.method(...)
    resolved_to: str                # best match: module:qualname OR "<unresolved>"
    resolved_confidence: str        # high/medium/low
    notes: str                      # extra info


# ----------------------------
# Helpers: AST unparse
# ----------------------------

def safe_unparse(node: ast.AST) -> str:
    if hasattr(ast, "unparse"):
        try:
            return ast.unparse(node)
        except Exception:
            return "<unparse_failed>"
    return "<unparse_unavailable>"


def format_arg(arg: ast.arg) -> str:
    if arg.annotation is not None:
        return f"{arg.arg}: {safe_unparse(arg.annotation)}"
    return arg.arg


def build_signature(fn: ast.AST) -> str:
    assert isinstance(fn, (ast.FunctionDef, ast.AsyncFunctionDef))
    a: ast.arguments = fn.args
    parts: List[str] = []

    posonly = getattr(a, "posonlyargs", [])
    for x in posonly:
        parts.append(format_arg(x))
    if posonly:
        parts.append("/")

    for x in a.args:
        parts.append(format_arg(x))

    if a.vararg is not None:
        parts.append("*" + format_arg(a.vararg))
    elif a.kwonlyargs:
        parts.append("*")

    for x, default in zip(a.kwonlyargs, a.kw_defaults):
        s = format_arg(x)
        if default is not None:
            s += f"={safe_unparse(default)}"
        parts.append(s)

    if a.kwarg is not None:
        parts.append("**" + format_arg(a.kwarg))

    # Defaults for normal args
    if a.defaults:
        rebuilt: List[str] = []
        for x in posonly:
            rebuilt.append(format_arg(x))
        if posonly:
            rebuilt.append("/")

        normal = a.args
        defaults = a.defaults
        cut = len(normal) - len(defaults)
        for idx, x in enumerate(normal):
            s = format_arg(x)
            if idx >= cut:
                s += f"={safe_unparse(defaults[idx - cut])}"
            rebuilt.append(s)

        if a.vararg is not None:
            rebuilt.append("*" + format_arg(a.vararg))
        elif a.kwonlyargs:
            rebuilt.append("*")

        for x, default in zip(a.kwonlyargs, a.kw_defaults):
            s = format_arg(x)
            if default is not None:
                s += f"={safe_unparse(default)}"
            rebuilt.append(s)

        if a.kwarg is not None:
            rebuilt.append("**" + format_arg(a.kwarg))

        parts = rebuilt

    ret = ""
    if fn.returns is not None:
        ret = f" -> {safe_unparse(fn.returns)}"

    return f"({', '.join(parts)}){ret}"


# ----------------------------
# Phase 1: Collect definitions
# ----------------------------

class DefVisitor(ast.NodeVisitor):
    def __init__(self, file_path: str, module_name: str) -> None:
        self.file_path = file_path
        self.module_name = module_name
        self.records: List[FunctionDefRecord] = []
        self.scope_stack: List[Tuple[str, str]] = [("<module>", "module")]  # (name, kind)

    def _current_scope_str(self) -> str:
        if len(self.scope_stack) == 1:
            return "<module>"
        return ".".join([n for (n, _) in self.scope_stack[1:]])

    def _in_class_scope(self) -> bool:
        return any(kind == "class" for _, kind in self.scope_stack)

    def _in_function_scope(self) -> bool:
        return any(kind == "function" for _, kind in self.scope_stack)

    def visit_ClassDef(self, node: ast.ClassDef) -> None:
        self.scope_stack.append((node.name, "class"))
        self.generic_visit(node)
        self.scope_stack.pop()

    def visit_FunctionDef(self, node: ast.FunctionDef) -> None:
        self._handle_function(node, is_async=False)
        self.scope_stack.append((node.name, "function"))
        self.generic_visit(node)
        self.scope_stack.pop()

    def visit_AsyncFunctionDef(self, node: ast.AsyncFunctionDef) -> None:
        self._handle_function(node, is_async=True)
        self.scope_stack.append((node.name, "function"))
        self.generic_visit(node)
        self.scope_stack.pop()

    def _handle_function(self, node: ast.AST, is_async: bool) -> None:
        assert isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))

        scope_str = self._current_scope_str()
        qual_parts = [] if scope_str == "<module>" else scope_str.split(".")
        qual_parts.append(node.name)
        qualname = ".".join(qual_parts)

        if self._in_class_scope() and not self._in_function_scope():
            def_type = "async_method" if is_async else "method"
        elif self._in_function_scope():
            def_type = "async_nested" if is_async else "nested"
        else:
            def_type = "async_function" if is_async else "function"

        signature = build_signature(node)
        decorators = ", ".join([safe_unparse(d) for d in node.decorator_list]) if node.decorator_list else ""
        lineno = getattr(node, "lineno", -1)
        end_lineno = getattr(node, "end_lineno", lineno)

        key = DefKey(module=self.module_name, qualname=qualname, def_type=def_type)

        self.records.append(
            FunctionDefRecord(
                key=key,
                file_path=self.file_path,
                module=self.module_name,
                scope=scope_str,
                name=node.name,
                qualname=qualname,
                def_type=def_type,
                signature=signature,
                decorators=decorators,
                lineno=lineno,
                end_lineno=end_lineno,
            )
        )


def module_name_from_path(root: Path, file_path: Path) -> str:
    rel = file_path.relative_to(root).with_suffix("")
    return ".".join(rel.parts)


def read_text_best_effort(file_path: Path) -> str:
    try:
        return file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return file_path.read_text(encoding="latin-1", errors="replace")


def parse_ast(file_path: Path) -> Optional[ast.AST]:
    src = read_text_best_effort(file_path)
    try:
        return ast.parse(src, filename=str(file_path))
    except SyntaxError:
        return None


# ----------------------------
# Phase 2: Collect imports + call sites
# ----------------------------

@dataclass
class ImportInfo:
    # local name -> module (import x as y => y -> x)
    imported_modules: Dict[str, str]
    # local name -> (module, original_name) for "from module import name as alias"
    imported_names: Dict[str, Tuple[str, str]]


class ImportVisitor(ast.NodeVisitor):
    def __init__(self) -> None:
        self.imported_modules: Dict[str, str] = {}
        self.imported_names: Dict[str, Tuple[str, str]] = {}

    def visit_Import(self, node: ast.Import) -> None:
        for alias in node.names:
            name = alias.name            # e.g. pkg.mod
            asname = alias.asname or name.split(".")[0]
            # store top-level binding
            self.imported_modules[asname] = name
        self.generic_visit(node)

    def visit_ImportFrom(self, node: ast.ImportFrom) -> None:
        if node.module is None:
            return
        mod = node.module
        for alias in node.names:
            if alias.name == "*":
                # can't statically map star imports well
                continue
            asname = alias.asname or alias.name
            self.imported_names[asname] = (mod, alias.name)
        self.generic_visit(node)


def get_call_target(call: ast.Call) -> Tuple[str, str]:
    """
    Return (kind, text):
      kind: "name" | "attr" | "other"
      text: best-effort representation of call.func
    """
    f = call.func
    if isinstance(f, ast.Name):
        return ("name", f.id)
    if isinstance(f, ast.Attribute):
        return ("attr", safe_unparse(f))
    return ("other", safe_unparse(f))


class UsageVisitor(ast.NodeVisitor):
    def __init__(
        self,
        file_path: str,
        module_name: str,
        import_info: ImportInfo,
        known_defs_by_name: Dict[str, List[FunctionDefRecord]],
        known_defs_by_module_attr: Dict[Tuple[str, str], List[FunctionDefRecord]],
        known_classes: Set[str],
    ) -> None:
        self.file_path = file_path
        self.module_name = module_name
        self.import_info = import_info
        self.known_defs_by_name = known_defs_by_name
        self.known_defs_by_module_attr = known_defs_by_module_attr
        self.known_classes = known_classes

        self.usages: List[UsageRecord] = []

    def visit_Call(self, node: ast.Call) -> None:
        kind, target_text = get_call_target(node)
        lineno = getattr(node, "lineno", -1)
        call_text = safe_unparse(node)

        resolved_to, confidence, notes = self.resolve_call(kind, target_text)
        self.usages.append(
            UsageRecord(
                used_in_file=self.file_path,
                used_in_module=self.module_name,
                used_in_lineno=lineno,
                call_text=call_text,
                resolved_to=resolved_to,
                resolved_confidence=confidence,
                notes=notes,
            )
        )

        self.generic_visit(node)

    def resolve_call(self, kind: str, target_text: str) -> Tuple[str, str, str]:
        """
        Try to resolve a call to a known definition.
        Returns (resolved_to, confidence, notes).
        resolved_to is "module:qualname" or "<unresolved>".
        """
        # Case 1: direct call foo()
        if kind == "name":
            local = target_text

            # from x import foo as bar => bar()
            if local in self.import_info.imported_names:
                mod, orig = self.import_info.imported_names[local]
                candidates = self.known_defs_by_module_attr.get((mod, orig), [])
                if candidates:
                    return (f"{candidates[0].module}:{candidates[0].qualname}", "high", "resolved via from-import")
                return ("<unresolved>", "low", f"from-import {mod}.{orig} not found in scanned defs")

            # local function in same module or elsewhere with same name
            candidates = self.known_defs_by_name.get(local, [])
            if candidates:
                # Prefer same module if available
                same_mod = [c for c in candidates if c.module == self.module_name]
                pick = same_mod[0] if same_mod else candidates[0]
                conf = "high" if same_mod else "medium"
                note = "resolved by name (same module preferred)" if same_mod else "resolved by name (ambiguous across modules)"
                return (f"{pick.module}:{pick.qualname}", conf, note)

            # maybe calling a class constructor (ClassName())
            if local in self.known_classes:
                return ("<unresolved>", "low", "looks like class constructor call; not a function def")

            return ("<unresolved>", "low", "no matching def by name")

        # Case 2: attribute call something.foo()
        if kind == "attr":
            # target_text is like "mod.foo" or "obj.foo" or "pkg.mod.foo"
            parts = target_text.split(".")
            if len(parts) >= 2:
                left = parts[0]
                attr = parts[-1]

                # import mod as m; m.foo()
                if left in self.import_info.imported_modules:
                    real_mod = self.import_info.imported_modules[left]
                    candidates = self.known_defs_by_module_attr.get((real_mod, attr), [])
                    if candidates:
                        return (f"{candidates[0].module}:{candidates[0].qualname}", "high", "resolved via imported module alias")
                    return ("<unresolved>", "low", f"module call {real_mod}.{attr} not found in scanned defs")

                # direct module usage without alias: mymod.foo()
                candidates = self.known_defs_by_module_attr.get((left, attr), [])
                if candidates:
                    return (f"{candidates[0].module}:{candidates[0].qualname}", "medium", "resolved via module.attr (module name guessed)")

                # class static call: MyClass.method()
                if left in self.known_classes:
                    # match any qualname ending with "MyClass.method"
                    wanted_suffix = f"{left}.{attr}"
                    matches = [c for c in self.known_defs_by_name.get(attr, []) if c.qualname.endswith(wanted_suffix)]
                    if matches:
                        return (f"{matches[0].module}:{matches[0].qualname}", "medium", "resolved as class method call")
                    return ("<unresolved>", "low", "looks like class method call but no matching method def found")

                # obj.method() is hard statically
                return ("<unresolved>", "low", "attribute call on unknown receiver (dynamic)")

            return ("<unresolved>", "low", "unhandled attribute shape")

        return ("<unresolved>", "low", "unhandled call type")


# ----------------------------
# Index building
# ----------------------------

def collect_class_names(tree: ast.AST) -> Set[str]:
    classes: Set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.ClassDef):
            classes.add(node.name)
    return classes


def build_indexes(defs: List[FunctionDefRecord]) -> Tuple[
    Dict[str, List[FunctionDefRecord]],
    Dict[Tuple[str, str], List[FunctionDefRecord]]
]:
    """
    known_defs_by_name: name -> defs
    known_defs_by_module_attr: (module, attr_name) -> defs (for resolving module.attr calls)
    """
    by_name: Dict[str, List[FunctionDefRecord]] = {}
    by_module_attr: Dict[Tuple[str, str], List[FunctionDefRecord]] = {}

    for d in defs:
        by_name.setdefault(d.name, []).append(d)

        # Provide module-level access mapping:
        # If a function is module-level "foo", map (module, "foo") -> it
        # If a method "MyClass.method", map (module, "method") too, but that can be ambiguous;
        # resolution code only uses this for module.attr calls.
        by_module_attr.setdefault((d.module, d.name), []).append(d)

    return by_name, by_module_attr


# ----------------------------
# Excel writing
# ----------------------------

def autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 90)


def write_excel(defs: List[FunctionDefRecord], usages: List[UsageRecord], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: Definitions
    ws1 = wb.active
    ws1.title = "Definitions"

    headers1 = [
        "File Path", "Module", "Scope", "Qualified Name", "Name", "Type",
        "Signature", "Decorators", "Line Start", "Line End"
    ]
    ws1.append(headers1)
    for i in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top")

    for d in sorted(defs, key=lambda x: (x.file_path, x.lineno, x.qualname)):
        ws1.append([
            d.file_path, d.module, d.scope, d.qualname, d.name, d.def_type,
            d.signature, d.decorators, d.lineno, d.end_lineno
        ])

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    autosize_columns(ws1)

    # Sheet 2: Usages
    ws2 = wb.create_sheet("Usages")
    headers2 = [
        "Used In File", "Used In Module", "Line", "Call Text",
        "Resolved To", "Confidence", "Notes"
    ]
    ws2.append(headers2)
    for i in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top")

    for u in sorted(usages, key=lambda x: (x.used_in_file, x.used_in_lineno)):
        ws2.append([
            u.used_in_file, u.used_in_module, u.used_in_lineno, u.call_text,
            u.resolved_to, u.resolved_confidence, u.notes
        ])

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    autosize_columns(ws2)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ----------------------------
# Main flow
# ----------------------------

def find_python_files(folder: Path) -> List[Path]:
    return [p for p in folder.rglob("*.py") if p.is_file()]


def main() -> int:
    # =========================
    # USER INPUT (EDIT THIS)
    # =========================
    FOLDER_PATH = "/localdata/user/kata_du/Automated Literature Survey/src"
    OUTPUT_EXCEL = "function_map.xlsx"
    # =========================

    root = Path(FOLDER_PATH).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        print(f"Error: folder not found or not a directory: {root}")
        return 2

    py_files = find_python_files(root)
    if not py_files:
        print(f"No .py files found under: {root}")
        return 0

    # Parse all ASTs once
    ast_by_file: Dict[Path, ast.AST] = {}
    module_by_file: Dict[Path, str] = {}
    class_names_global: Set[str] = set()

    for f in py_files:
        tree = parse_ast(f)
        if tree is None:
            continue
        ast_by_file[f] = tree
        mod = module_name_from_path(root, f)
        module_by_file[f] = mod
        class_names_global |= collect_class_names(tree)

    # Collect definitions
    defs: List[FunctionDefRecord] = []
    for f, tree in ast_by_file.items():
        visitor = DefVisitor(str(f), module_by_file[f])
        visitor.visit(tree)
        defs.extend(visitor.records)

    known_defs_by_name, known_defs_by_module_attr = build_indexes(defs)

    # Collect usages
    usages: List[UsageRecord] = []
    for f, tree in ast_by_file.items():
        iv = ImportVisitor()
        iv.visit(tree)
        import_info = ImportInfo(
            imported_modules=iv.imported_modules,
            imported_names=iv.imported_names,
        )

        uv = UsageVisitor(
            file_path=str(f),
            module_name=module_by_file[f],
            import_info=import_info,
            known_defs_by_name=known_defs_by_name,
            known_defs_by_module_attr=known_defs_by_module_attr,
            known_classes=class_names_global,
        )
        uv.visit(tree)
        usages.extend(uv.usages)

    out_path = Path(OUTPUT_EXCEL).expanduser().resolve()
    write_excel(defs, usages, out_path)

    print(f"Scanned {len(py_files)} Python files.")
    print(f"Found {len(defs)} function/method definitions.")
    print(f"Found {len(usages)} call-sites (some unresolved).")
    print(f"Excel written to: {out_path}")
    return 0



if __name__ == "__main__":
    raise SystemExit(main())
